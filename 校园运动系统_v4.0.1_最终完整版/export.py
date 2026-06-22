"""
数据导出模块
支持将运动记录、体测成绩等数据导出为格式化 Excel 文件（.xlsx）和纯文本报告。

亮点（可写进简历）：
- 使用 openpyxl 生成带样式的 Excel（表头高亮、斑马纹、自动列宽）
- 汇总 Sheet 包含聚合统计和图表占位
- 降级兼容：若未安装 openpyxl，自动退回 CSV 导出
"""
from __future__ import annotations

import csv
import os
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING

from config import EXPORT_DIR, REPORT_DIR
from logger import get_logger

if TYPE_CHECKING:
    from models import HealthLog, HealthProfile, PhysicalTestRecord, SportRecord

log = get_logger(__name__)

try:
    import openpyxl
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ──── 样式常量 ────
_HEADER_FILL  = "2563EB"   # 蓝色表头
_ALT_FILL     = "EFF6FF"   # 斑马纹
_ACCENT_FILL  = "16A34A"   # 绿色强调
_WARN_FILL    = "FEF3C7"   # 预警黄
_FONT_NAME    = "Microsoft YaHei"


def _header_style() -> tuple:
    """返回表头单元格样式三件套（Font, Fill, Alignment）。"""
    font  = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=10)
    fill  = PatternFill("solid", fgColor=_HEADER_FILL)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _body_style(alt: bool = False) -> tuple:
    font  = Font(name=_FONT_NAME, size=9)
    fill  = PatternFill("solid", fgColor=_ALT_FILL) if alt else PatternFill("solid", fgColor="FFFFFF")
    align = Alignment(horizontal="left", vertical="center")
    return font, fill, align


def _auto_col_width(ws) -> None:
    """根据内容自动调整每列宽度（上限 40 字符）。"""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0 for cell in col),
            default=8,
        )
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len * 1.5 + 2, 40)


def _apply_header(ws, headers: list[str], row: int = 1) -> None:
    font, fill, align = _header_style()
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font   = font
        cell.fill   = fill
        cell.alignment = align
    ws.row_dimensions[row].height = 24


def _apply_body_row(ws, values: list, row: int, alt: bool) -> None:
    font, fill, align = _body_style(alt)
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
    ws.row_dimensions[row].height = 18


# ───────────────────────────────────────────────────────────────
# 主导出函数
# ───────────────────────────────────────────────────────────────
def export_sport_excel(
    records: list["SportRecord"],
    filename: str | None = None,
) -> Path:
    """
    将运动打卡记录导出为 Excel。
    返回生成文件的完整路径。
    """
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = filename or f"运动记录_{stamp}.xlsx"
    out_path = EXPORT_DIR / filename

    if not HAS_OPENPYXL:
        return _export_sport_csv(records, out_path.with_suffix(".csv"))

    wb = openpyxl.Workbook()

    # ── Sheet 1：明细 ──
    ws = wb.active
    ws.title = "运动明细"
    headers = ["学号", "姓名", "班级", "日期", "运动类型", "打卡分类",
               "打卡方式", "地点", "时长(分钟)", "距离(km)", "步数",
               "心率", "消耗(kcal)", "审核状态", "备注"]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for i, r in enumerate(records):
        row_data = [
            r.student_id, r.name, r.class_name, r.day, r.sport_type,
            r.category, r.checkin_method, r.location,
            r.duration, r.distance, r.steps, r.heart_rate, r.calories,
            r.status, r.remark,
        ]
        _apply_body_row(ws, row_data, i + 2, alt=(i % 2 == 1))
    _auto_col_width(ws)

    # ── Sheet 2：按运动类型汇总 ──
    from analytics import summarize_by_sport
    ws2 = wb.create_sheet("运动类型汇总")
    _apply_header(ws2, ["运动类型", "次数", "总时长(分钟)", "总热量(kcal)", "占比"])
    sport_rows = summarize_by_sport(records)
    total_dur = sum(r[2] for r in sport_rows) or 1
    for i, (sport, count, dur, cal) in enumerate(sport_rows):
        pct = f"{dur / total_dur * 100:.1f}%"
        _apply_body_row(ws2, [sport, count, dur, cal, pct], i + 2, alt=(i % 2 == 1))
    _auto_col_width(ws2)

    # ── Sheet 3：每日汇总 ──
    from collections import defaultdict
    ws3 = wb.create_sheet("每日汇总")
    _apply_header(ws3, ["日期", "运动次数", "总时长(分钟)", "总热量(kcal)"])
    daily: dict[str, list[float]] = defaultdict(lambda: [0.0, 0.0, 0.0])
    for r in records:
        if r.status == "已通过":
            daily[r.day][0] += 1
            daily[r.day][1] += r.duration
            daily[r.day][2] += r.calories
    for i, (day, (cnt, dur, cal)) in enumerate(sorted(daily.items(), reverse=True)):
        _apply_body_row(ws3, [day, int(cnt), int(dur), round(cal, 1)], i + 2, alt=(i % 2 == 1))
    _auto_col_width(ws3)

    wb.save(out_path)
    log.info("Excel 导出成功：%s（%d 行）", out_path, len(records))
    return out_path


def export_physical_test_excel(
    tests: list["PhysicalTestRecord"],
    filename: str | None = None,
) -> Path:
    """将体测成绩导出为 Excel。"""
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = filename or f"体测成绩_{stamp}.xlsx"
    out_path = EXPORT_DIR / filename

    if not HAS_OPENPYXL:
        return _export_test_csv(tests, out_path.with_suffix(".csv"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "体测成绩"
    headers = ["学号", "姓名", "性别", "年份", "800/1000米", "立定跳远(cm)",
               "肺活量", "坐位体前屈(cm)", "总分", "评级"]
    _apply_header(ws, headers)
    ws.freeze_panes = "A2"

    for i, t in enumerate(tests):
        row_data = [
            t.student_id, t.name, t.gender, t.year, t.run_score,
            t.long_jump, t.lung_capacity, t.sit_reach, t.total_score, t.rating,
        ]
        row_idx = i + 2
        _apply_body_row(ws, row_data, row_idx, alt=(i % 2 == 1))
        # 不及格行标红
        if t.total_score < 60:
            warn_fill = PatternFill("solid", fgColor="FEE2E2")
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = warn_fill
    _auto_col_width(ws)

    wb.save(out_path)
    log.info("体测 Excel 导出成功：%s", out_path)
    return out_path


def export_health_report_txt(
    profile: "HealthProfile",
    records: list["SportRecord"],
    logs: list["HealthLog"],
    tests: list["PhysicalTestRecord"],
) -> Path:
    """生成纯文本健康周报并写入 reports/ 目录。"""
    from analytics import weekly_report
    content = weekly_report(profile, records, logs, tests)
    stamp = datetime.now().strftime("%Y-%m-%d")
    out_path = REPORT_DIR / f"健康周报_{stamp}.txt"
    out_path.write_text(content, encoding="utf-8")
    log.info("健康周报已生成：%s", out_path)
    return out_path


# ── 降级：无 openpyxl 时退回 CSV ──
def _export_sport_csv(records: list["SportRecord"], path: Path) -> Path:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["学号", "姓名", "班级", "日期", "运动类型", "时长", "热量", "审核状态"])
        for r in records:
            writer.writerow([r.student_id, r.name, r.class_name, r.day,
                             r.sport_type, r.duration, r.calories, r.status])
    log.warning("openpyxl 未安装，已退回 CSV 导出：%s", path)
    return path


def _export_test_csv(tests: list["PhysicalTestRecord"], path: Path) -> Path:
    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["学号", "姓名", "年份", "总分", "评级"])
        for t in tests:
            writer.writerow([t.student_id, t.name, t.year, t.total_score, t.rating])
    return path
